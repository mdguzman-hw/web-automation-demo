# Copyright © 2026 - Homewood Health Inc.

# Pytest configuration and shared fixtures
import os
from datetime import datetime

import pytest
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from suites.CustomerPortal import CustomerPortal
from suites.Homeweb import Homeweb
from suites.QuantumAPI import QuantumAPI
from suites.SentioClient import SentioClient
from suites.SentioProvider import SentioProvider

# --- Report Collection ---

_run_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
_date_str = _run_timestamp[4:6] + "-" + _run_timestamp[6:8] + "-" + _run_timestamp[:4]
_reports_dir = f"reports/{_date_str}/{_run_timestamp}"
_env_results = {}
_versions = {}  # {"PROD": {"Homeweb": "v3.0.17.261", ...}, "BETA": {...}}
_pending_output = {}  # {nodeid: [messages...]}
_all_results = []  # [(nodeid, phase, skip_reason, stdout), ...]
_test_metadata = {}  # {func_name: {"id": "BAT-WEB-001", "description": "...", "suite": "..."}}


def _pct(r):
    total = r["passed"] + r["failed"]
    return f"{int(r['passed'] / total * 100)}%" if total > 0 else "N/A"


@pytest.fixture(scope="function")
def record_output(request):
    nodeid = request.node.nodeid

    def _record(message):
        _pending_output.setdefault(nodeid, []).append(message)
        print(message)

    return _record


@pytest.fixture(scope="session")
def record_version():
    import re
    import requests

    def _record(label, base_url, env):
        try:
            r = requests.get(f"{base_url}/version.html", timeout=5, verify=False)
            match = re.search(r"v?\d+\.\d+\.\d+\.\d+", r.text)
            raw = match.group(0) if match else "N/A"
            version = raw if raw == "N/A" or raw.startswith("v") else f"v{raw}"
        except Exception:
            version = "N/A"
        env_key = env.upper()
        if env_key not in _versions:
            _versions[env_key] = {}
        _versions[env_key][label] = version

    return _record


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    import re
    outcome = yield
    rep = outcome.get_result()
    if rep.when == "call" and rep.failed:
        driver = item.funcargs.get("driver")
        if driver:
            os.makedirs(_reports_dir, exist_ok=True)
            name_match = re.search(r"::(\w+)", rep.nodeid)
            test_name = name_match.group(1).replace("test_", "") if name_match else "unknown"
            env_match = re.search(r"\[(\w+)\]", rep.nodeid)
            env_tag = env_match.group(1) if env_match else ""
            path = f"{_reports_dir}/fail-{test_name}-{env_tag}_{_run_timestamp}.png"
            driver.save_screenshot(path)
            print(f"\nScreenshot saved: {path}")


def pytest_runtest_logreport(report):
    if report.when == "setup" and report.skipped:
        phase = "skipped"
        reason = report.longrepr[2] if isinstance(report.longrepr, tuple) else str(report.longrepr)
        reason = reason.replace("Skipped: ", "")
        _all_results.append((report.nodeid, "SKIPPED", reason, ""))
    elif report.when == "call":
        if report.passed:
            phase = "passed"
            stdout = "\n".join(_pending_output.pop(report.nodeid, []))
            _all_results.append((report.nodeid, "PASSED", "", stdout))
        elif report.failed:
            phase = "failed"
            stdout = "\n".join(_pending_output.pop(report.nodeid, []))
            longrepr = str(report.longrepr).strip() if report.longrepr else ""
            _all_results.append((report.nodeid, "FAILED", longrepr, stdout))
        elif report.skipped:
            phase = "skipped"
            reason = report.longrepr[2] if isinstance(report.longrepr, tuple) else str(report.longrepr)
            reason = reason.replace("Skipped: ", "")
            _all_results.append((report.nodeid, "SKIPPED", reason, ""))
        else:
            return
    else:
        return

    name = report.nodeid
    if "[PROD]" in name:
        env = "PROD"
    elif "[BETA]" in name:
        env = "BETA"
    elif "[LOCAL]" in name:
        env = "LOCAL"
    else:
        env = "PROD"

    if env not in _env_results:
        _env_results[env] = {"passed": 0, "failed": 0, "skipped": 0}

    _env_results[env][phase] += 1


def pytest_collection_finish(session):
    import re
    import inspect

    suite_map = {
        "bat_homeweb-new": "Homeweb - New",
        "bat_homeweb":     "Homeweb",
        "bat_web":         "Web",
        "smoke_homeweb":   "Homeweb Smoke",
        "smoke_web":       "Web Smoke",
    }

    for item in session.items:
        func_name = item.function.__name__
        if func_name in _test_metadata:
            continue
        try:
            source_file = inspect.getsourcefile(item.function)
            start_line = item.function.__code__.co_firstlineno
            with open(source_file) as f:
                all_lines = f.readlines()
        except Exception:
            continue

        bat_id = description = None
        for i in range(start_line - 2, max(start_line - 6, -1), -1):
            m = re.match(r"#\s*((?:BAT|SMOKE)-[\w-]+)\s*\|\s*(.+)", all_lines[i].strip())
            if m:
                bat_id, description = m.group(1), m.group(2).strip()
                break

        if not bat_id:
            continue

        stem = os.path.basename(source_file).replace(".py", "").replace("test_", "")
        _test_metadata[func_name] = {
            "id": bat_id,
            "description": description,
            "suite": suite_map.get(stem, stem),
        }


def _write_test_matrix_xlsx(report_name, active_envs, run_time="-"):
    import re
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # --- build per-test lookup ---
    test_phases = {}
    test_stdout = {}
    test_detail = {}
    for nodeid, phase, detail, stdout in _all_results:
        m = re.search(r"::(\w+)(?:\[(\w+)\])?", nodeid)
        if not m:
            continue
        func_name = m.group(1)
        env = m.group(2) if m.group(2) else "PROD"
        test_phases.setdefault(func_name, {})[env] = phase
        test_stdout.setdefault(func_name, {})[env] = stdout
        test_detail.setdefault(func_name, {})[env] = detail

    ordered = sorted(
        [(meta["id"], fn, meta) for fn, meta in _test_metadata.items()],
        key=lambda x: x[0]
    )

    # --- styles ---
    header_font = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    right_bold  = Alignment(horizontal="right")
    grey_fill   = PatternFill("solid", fgColor="D9D9D9")
    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = header_font
        c.fill = grey_fill
        c.alignment = center
        c.border = border
        return c

    def data_cell(row, col, value, align=None):
        c = ws.cell(row=row, column=col, value=value)
        c.alignment = align or left_wrap
        c.border = border
        return c

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Matrix"

    # Columns: A=ID, B=Description, C=Logs, then one per env
    # TOTALS labels also go in col C, counts in the env columns
    LOGS_COL = 3
    env_cols = {env: LOGS_COL + 1 + i for i, env in enumerate(active_envs)}

    # --- column headers (row 1) ---
    hdr_cell(1, 1, "ID")
    hdr_cell(1, 2, "Description")
    hdr_cell(1, LOGS_COL, "Logs")
    for env, col in env_cols.items():
        hdr_cell(1, col, env)

    # --- column widths ---
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 55
    for col in env_cols.values():
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12

    # --- data rows ---
    counts = {env: {"passed": 0, "failed": 0, "skipped": 0} for env in active_envs}

    for offset, (bat_id, func_name, meta) in enumerate(ordered):
        row = 2 + offset
        data_cell(row, 1, meta["id"])
        data_cell(row, 2, meta["description"])

        # Logs: stdout from any env; append failure detail for failed envs
        logs = next(
            (test_stdout.get(func_name, {}).get(env, "").strip()
             for env in active_envs if test_stdout.get(func_name, {}).get(env, "").strip()),
            ""
        )
        for env in active_envs:
            if test_phases.get(func_name, {}).get(env) == "FAILED":
                detail = test_detail.get(func_name, {}).get(env, "")
                last_line = next(
                    (l.strip() for l in reversed(detail.splitlines()) if l.strip()), ""
                ) if detail else ""
                if last_line:
                    logs = (logs + "\n" if logs else "") + f"[{env} FAILED] {last_line[:120]}"
        c = data_cell(row, LOGS_COL, logs or "-")
        if logs:
            ws.row_dimensions[row].height = max(15, logs.count("\n") * 15 + 15)

        for env, col in env_cols.items():
            phase = test_phases.get(func_name, {}).get(env)
            if phase == "PASSED":
                symbol, fill = "Y", green_fill
                counts[env]["passed"] += 1
            elif phase == "FAILED":
                symbol, fill = "N", red_fill
                counts[env]["failed"] += 1
            else:
                symbol, fill = "-", None
                counts[env]["skipped"] += 1
            c = data_cell(row, col, symbol, align=center)
            if fill:
                c.fill = fill

    # --- totals (labels in col C, counts in env cols) ---
    totals_row = 2 + len(ordered) + 1

    # TOTALS header — env names in the result columns
    c = ws.cell(row=totals_row, column=LOGS_COL, value="TOTALS")
    c.font = header_font
    c.alignment = right_bold
    for env, col in env_cols.items():
        ws.cell(row=totals_row, column=col, value=env).alignment = center

    # Build row
    totals_row += 1
    c = ws.cell(row=totals_row, column=LOGS_COL, value="Build")
    c.font = header_font
    c.alignment = right_bold
    for env, col in env_cols.items():
        version_lines = "\n".join(f"{k}: {v}" for k, v in _versions.get(env, {}).items())
        c = ws.cell(row=totals_row, column=col, value=version_lines or "-")
        c.alignment = left_wrap
        if version_lines:
            ws.row_dimensions[totals_row].height = max(15, version_lines.count("\n") * 15 + 15)

    for label, key in [
        ("Passed",                       "passed"),
        ("Failed",                       "failed"),
        ("Not Run (Skipped, N/A, etc.)", "skipped"),
        ("Completed",                    "passed"),
    ]:
        totals_row += 1
        c = ws.cell(row=totals_row, column=LOGS_COL, value=label)
        c.font = header_font
        c.alignment = right_bold
        for env, col in env_cols.items():
            ws.cell(row=totals_row, column=col, value=counts[env][key]).alignment = center

    totals_row += 1
    c = ws.cell(row=totals_row, column=LOGS_COL, value="Percentage Passed")
    c.font = header_font
    c.alignment = right_bold
    for env, col in env_cols.items():
        total = counts[env]["passed"] + counts[env]["failed"]
        pct = f"{int(counts[env]['passed'] / total * 100)}%" if total > 0 else "N/A"
        ws.cell(row=totals_row, column=col, value=pct).alignment = center

    totals_row += 1
    c = ws.cell(row=totals_row, column=LOGS_COL, value="Run Time")
    c.font = header_font
    c.alignment = right_bold
    first_env_col = min(env_cols.values())
    ws.cell(row=totals_row, column=first_env_col, value=run_time).alignment = center

    path = f"{_reports_dir}/{_run_timestamp}_{report_name}_report.xlsx"
    wb.save(path)
    return path


def pytest_terminal_summary(terminalreporter, exitstatus, config):
    if not _env_results:
        return

    envs = ["PROD", "BETA", "LOCAL"]
    results = {e: _env_results.get(e, {"passed": 0, "failed": 0, "skipped": 0}) for e in envs}

    total_passed = sum(r["passed"] for r in results.values())
    total_failed = sum(r["failed"] for r in results.values())
    total_skipped = sum(r["skipped"] for r in results.values())
    total_run = total_passed + total_failed
    total_pct = f"{int(total_passed / total_run * 100)}%" if total_run > 0 else "N/A"

    duration = getattr(terminalreporter, "_sessionduration", None)
    run_time = f"{int(duration // 60)}m {int(duration % 60)}s" if duration is not None else "N/A"

    sep = "-" * 62
    terminalreporter.write_sep("=", "TEST REPORT")

    if _versions:
        prod_lines  = [f"{k}: {v}" for k, v in _versions.get("PROD",  {}).items()]
        beta_lines  = [f"{k}: {v}" for k, v in _versions.get("BETA",  {}).items()]
        local_lines = [f"{k}: {v}" for k, v in _versions.get("LOCAL", {}).items()]
        for i in range(max(len(prod_lines), len(beta_lines), len(local_lines))):
            p = prod_lines[i]  if i < len(prod_lines)  else ""
            b = beta_lines[i]  if i < len(beta_lines)  else ""
            l = local_lines[i] if i < len(local_lines) else ""
            terminalreporter.write_line(f"  {p:<33} {b:<20} {l}")
        terminalreporter.write_line("")

    terminalreporter.write_line(f"{'Metric':<35} {'PROD':>7} {'BETA':>7} {'LOCAL':>7}")
    terminalreporter.write_line(sep)
    terminalreporter.write_line(f"{'Passed':<35} {results['PROD']['passed']:>7} {results['BETA']['passed']:>7} {results['LOCAL']['passed']:>7}")
    terminalreporter.write_line(f"{'Failed':<35} {results['PROD']['failed']:>7} {results['BETA']['failed']:>7} {results['LOCAL']['failed']:>7}")
    terminalreporter.write_line(f"{'Not Run (Skipped, N/A, etc.)':<35} {results['PROD']['skipped']:>7} {results['BETA']['skipped']:>7} {results['LOCAL']['skipped']:>7}")
    terminalreporter.write_line(f"{'Completed':<35} {results['PROD']['passed'] + results['PROD']['failed']:>7} {results['BETA']['passed'] + results['BETA']['failed']:>7} {results['LOCAL']['passed'] + results['LOCAL']['failed']:>7}")
    terminalreporter.write_line(f"{'Percentage Passed':<35} {_pct(results['PROD']):>7} {_pct(results['BETA']):>7} {_pct(results['LOCAL']):>7}")
    terminalreporter.write_line(sep)
    terminalreporter.write_line(f"{'Total Passed':<35} {total_passed:>7}")
    terminalreporter.write_line(f"{'Total Failed':<35} {total_failed:>7}")
    terminalreporter.write_line(f"{'Total Not Run':<35} {total_skipped:>7}")
    terminalreporter.write_line(f"{'Total Percentage Passed':<35} {total_pct:>7}")

    file_args = [a for a in config.invocation_params.args if a.endswith(".py")]
    if file_args:
        stem = os.path.basename(file_args[0]).replace(".py", "")
        report_name = stem.replace("test_", "")
    else:
        report_name = "report"

    if _test_metadata:
        os.makedirs(_reports_dir, exist_ok=True)
        active_envs = [e for e in ["PROD", "BETA", "LOCAL"] if e in _env_results]
        matrix_path = _write_test_matrix_xlsx(report_name, active_envs, run_time)
        short_path = matrix_path.replace("reports/", "/")
        terminalreporter.write_line(f"\n  Report saved: {short_path}")


@pytest.fixture(scope="session")
def driver(env):
    # 1: Configure Chrome options
    chrome_options = Options()
    # chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--incognito")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")

    # 2: Launch Browser
    driver_instance = webdriver.Chrome(options=chrome_options)

    # 3. YIELD - give driver to the tests
    yield driver_instance

    # 4: Close Browser
    driver_instance.quit()


def pytest_addoption(parser):
    parser.addoption(
        "--env",
        action="store",
        default="all",
        help="Environment: prod | beta | local | all (default: all)"
    )


def pytest_collection_modifyitems(items):
    def get_group(item):
        name = item.name

        import re
        match = re.search(r"test_bat_web_(\d+)", name)
        order = int(match.group(1)) if match else 999

        # 🔥 ENV FIRST (this is the fix)
        if "[PROD]" in name:
            env = 0
        elif "[BETA]" in name:
            env = 1
        elif "[LOCAL]" in name:
            env = 2
        elif "_beta" in name.lower():
            env = 1
        else:
            env = 3

        return env, order

    items.sort(key=get_group)


@pytest.fixture(params=["prod", "beta", "local"], ids=["PROD", "BETA", "LOCAL"], scope="session")
def env(request):
    env_flag = request.config.getoption("--env")

    if env_flag != "all" and request.param != env_flag:
        pytest.skip(f"Skipping {request.param} environment")

    return request.param


load_dotenv()


@pytest.fixture(scope="session")
def credentials():
    return {
        "hhi_s1": {
            "email": os.getenv("HHI_S1_EMAIL"),
            "password": os.getenv("HHI_S1_PASSWORD")
        },
        "hhi_personal": {
            "email": os.getenv("HHI_PERSONAL_EMAIL"),
            "password": os.getenv("HHI_PERSONAL_PASSWORD")
        },
        "dsg_demo": {
            "email": os.getenv("DSG_DEMO_EMAIL"),
            "password": os.getenv("DSG_DEMO_PASSWORD")
        },
        "sentio": {
            "email": os.getenv("SENTIO_EMAIL"),
            "password": os.getenv("SENTIO_PASSWORD")
        },
        "hhi_demo": {
            "email": os.getenv("HHI_DEMO_EMAIL"),
            "password": os.getenv("HHI_DEMO_PASSWORD")
        },
        "lso_test": {
            "email": os.getenv("LSO_EMAIL"),
            "password": os.getenv("LSO_PASSWORD")
        },
        "fresh_1": {
            "email": os.getenv("FRESH_1_EMAIL"),
            "password": os.getenv("FRESH_1_PASSWORD")
        },
        "dsgdemo_s2": {
            "email": os.getenv("DSGDEMO_S2_EMAIL"),
            "password": os.getenv("DSGDEMO_S2_PASSWORD")
        },
        "dsgdemo_s3": {
            "email": os.getenv("DSGDEMO_S3_EMAIL"),
            "password": os.getenv("DSGDEMO_S3_PASSWORD")
        }
    }


@pytest.fixture(scope="session")
def language():
    return os.getenv("LANGUAGE", "en")


@pytest.fixture(scope="session")
def quantum(driver, language, env):
    return QuantumAPI(driver, language, env)


@pytest.fixture(scope="session")
def homeweb(driver, language, env, quantum):
    return Homeweb(driver, language, env, quantum)


@pytest.fixture(scope="session")
def customer_portal(driver, language, env, quantum):
    if env == "beta":
        return pytest.skip(f"Skipping {env} environment")
    else:
        return CustomerPortal(driver, language, env, quantum)


@pytest.fixture(scope="session")
def quantum_prod(driver, language):
    return QuantumAPI(driver, language, "prod")


@pytest.fixture(scope="session")
def sentio_client(driver, language, env, quantum, quantum_prod):
    if env == "prod":
        return pytest.skip(f"Skipping {env} environment")
    else:
        return SentioClient(driver, language, env, quantum_prod)


@pytest.fixture(scope="session")
def sentio_provider(driver, language, env, quantum, quantum_prod):
    if env == "prod":
        return pytest.skip(f"Skipping {env} environment")
    else:
        return SentioProvider(driver, language, env, quantum_prod)


@pytest.fixture(scope="function")
def record_account():
    import openpyxl

    HEADERS = ["Date", "Timestamp", "Env", "Org", "Division", "Role",
               "First Name", "Last Name", "Preferred Name", "Email", "DOB", "Password", "Marketing Opt-in"]
    ACCOUNTS_PATH = "reports/registered-accounts.xlsx"

    def _record(env, org, division, role, first_name, last_name, preferred_name, email, dob, password, marketing_opt_in):
        os.makedirs("reports", exist_ok=True)
        if os.path.exists(ACCOUNTS_PATH):
            wb = openpyxl.load_workbook(ACCOUNTS_PATH)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Registered Accounts"
            ws.append(HEADERS)

        ws.append([
            _run_timestamp[4:6] + "-" + _run_timestamp[6:8] + "-" + _run_timestamp[:4],
            _run_timestamp,
            env.upper(),
            org,
            division,
            role,
            first_name,
            last_name,
            preferred_name,
            email,
            dob,
            password,
            "Yes" if marketing_opt_in else "No",
        ])
        wb.save(ACCOUNTS_PATH)

    return _record


@pytest.fixture(scope="session")
def latest_registered_account():
    import openpyxl

    ACCOUNTS_PATH = "reports/registered-accounts.xlsx"
    HEADERS = ["Date", "Timestamp", "Env", "Org", "Division", "Role",
               "First Name", "Last Name", "Preferred Name", "Email", "DOB", "Password", "Marketing Opt-in"]

    if not os.path.exists(ACCOUNTS_PATH):
        pytest.skip("No registered accounts found — run test_bat_web_003 first")

    wb = openpyxl.load_workbook(ACCOUNTS_PATH)
    ws = wb.active

    if ws.max_row < 2:
        pytest.skip("No registered accounts found — run test_bat_web_003 first")

    values = [ws.cell(row=ws.max_row, column=i + 1).value for i in range(len(HEADERS))]
    return dict(zip(HEADERS, values))
